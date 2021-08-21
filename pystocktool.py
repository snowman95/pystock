import sys, subprocess, tickers, re, datetime
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from itertools import islice
try:
    import requests
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'requests'])
finally:
    import requests
try:
    import finpie
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", 'finpie'])
finally:
    import finpie
from pdb import set_trace as pb
if sys.version_info[0] < 3: 
    from StringIO import StringIO
else:
    from io import StringIO


### Option Setting
pd.set_option('display.max_row', None)
pd.set_option('display.max_columns', None)  # 터미널 화면에 축약없이 전체내용 출력
pd.set_option('display.max_colwidth', None)
pd.set_option('display.width', None)
pd.set_option('display.date_yearfirst',True)

billion = 1000000000.0 # 10억(B)
excel_path = 'C:\\Users\\hoon\\SDS\\stock_report\\'
excel_sheet_name = 'Portfolio'

time_now = datetime.datetime.today()
time_today = time_now.date()
portfolio_file_name = 'Portfolio.xlsx'
portfolio_file_sheet_title = 'Portfolio'
report_file_name = str(time_today) + '_' + 'Stock_Report.xlsx'
new_report_file_name = f'{str(time_today)} {time_now.hour}-{time_now.minute}-{time_now.second}_Stock_Report.xlsx'
report_file_sheet_title = 'Report'

# 회사이름은 https://www.macrotrends.net/stocks/stock-screener 의 html 소스보기 - originalData 딕셔너리를 가져오면 됨.

#	revenue_ttm = stock_datas_f['sales']
#	shares_outstanding = stock_datas_y['shares_outstanding']

def converted_to_billions(price):
	if(isinstance(price, str)):
		price = float(price)
	result = round(price/billion,3)
	return result

# 현재-days ~ 현재 데이터 가져옴
def get_historical_price_data(ticker, before_days=365):
	now = datetime.datetime.now()
	start = now + datetime.timedelta(days=-before_days)
	start_date = int(start.timestamp())
	end_date   = int(now.timestamp())

	crumble_link = f'https://finance.yahoo.com/quote/{ticker}/history?p={ticker}'
	crumble_regex = r'CrumbStore":{"crumb":"(.*?)"}'
	#cookie_regex = r'set-cookie: (.*?);'
	#quote_link = 'https://query1.finance.yahoo.com/v7/finance/download/{}?period1={}&period2={}&interval=1d&events=history&crumb={}'
	
	link = crumble_link.format(ticker)
	print(link)
	session = requests.Session()
	response = session.get(link)
	# get crumbs
	text = str(response.content)
	match = re.search(crumble_regex, text)
	crumbs = match.group(1)
	
	# get cookie
	cookie = session.cookies.get_dict()
	url = f"https://query1.finance.yahoo.com/v7/finance/download/{ticker}?period1={start_date}&period2={end_date}&interval=1d&events=history&crumb={crumbs}"
	#r = requests.get(url,cookies=session.cookies.get_dict(),timeout=5, stream=True)
	try:
		r = requests.get(url,cookies=cookie,timeout=10, stream=True)
	except:
		print('[Error] get_historical_price_data() 실패 - timeout 늘려서 재시도')
		try:
			r = requests.get(url,cookies=cookie,timeout=15, stream=True)
		except : 
			pass 
	input = StringIO(r.text)
	df = pd.read_csv(StringIO(r.text), sep=",", header=0)# index_col=0)
	df.index = pd.to_datetime(df['Date'])
	return df
	
# output : ['high':high, 'low':low, 'transition':transition, 'trends_days':trends_days]
def trends_anlysis(si, standard, acceptable_transition_day):
	high = 0
	low = 0
	transition = 0
	trends_days = 0
	current = si[0]
	plus_trend_si  = si[si >=standard] if si[si >=standard].any() else si.max()
	minus_trend_si = si[si < standard] if si[si < standard].any() else si.min()

	# Case 1. 양/음 영역 동시 존재
	if isinstance(plus_trend_si, pd.Series) == True and isinstance(minus_trend_si, pd.Series) == True :
		transition = '동일'
		if current >= standard :
			reversed_day = minus_trend_si.index[0].date()  # 음 → 양의 영역 역전되는 마지막 날
			last_plus_trend_si = plus_trend_si.loc[:reversed_day] # 역전된 날 ~ 현재
			high = last_plus_trend_si.max()
			low = last_plus_trend_si.min()
			trends_days = len(last_plus_trend_si)
			if len(last_plus_trend_si) <= acceptable_transition_day :
				transition = '음→양'
		else :
			reversed_day = plus_trend_si.index[0].date()
			last_minus_trend_si = minus_trend_si.loc[:reversed_day]
			high = last_minus_trend_si.max()
			low = last_minus_trend_si.min()
			trends_days = len(last_minus_trend_si)
			if len(last_minus_trend_si) <= acceptable_transition_day :
				transition = '양→음'
	# Case 2. 양의 영역만 존재
	elif isinstance(minus_trend_si, pd.Series) == False :
		high = plus_trend_si.max()
		low = plus_trend_si.min()
		trends_days = len(plus_trend_si)
	# Case 3. 음의 영역만 존재
	elif isinstance(plus_trend_si, pd.Series) == False :
		high = minus_trend_si.max()
		low = minus_trend_si.min()
		trends_days = len(minus_trend_si)

	return {'high':high, 'low':low, 'transition':transition, 'trends_days':trends_days}

# return : 추세전환(음→양)/매수세(X일)/과매수(Y일)
def get_rsi_trends(price_si, dist_from_52w_high, dist_from_52w_low, days=180, acceptable_transition_day=2):
	overbought = 70; standard = 50; oversold = 30
	price_si = price_si[-days:]
	delta = price_si.diff(1)
	delta = delta.dropna()
	up = delta.copy()
	down = delta.copy()
	up[up<0] = 0
	down[down>0] = 0
	avg_gain = up.rolling(window=14).mean()
	avg_loss = abs(down.rolling(window=14).mean())
	rsi_si = 100.0 - (100.0/(1.0+avg_gain/avg_loss))
	rsi_si = rsi_si.iloc[::-1] # 행 순서 반대로
	rsi_current = rsi_si[0]
	trend_dict = trends_anlysis(rsi_si, standard, acceptable_transition_day)

	if rsi_current > overbought :
		trend = '[과]매수'
		over_trends_days = trends_anlysis(rsi_si,overbought,1)['trends_days']
		trend_dict['trends_days'] = f'[{over_trends_days}일]{trend_dict["trends_days"]}'
	elif rsi_current >= standard :
		trend = '매수'
	elif rsi_current < oversold :
		trend = '[과]매도'
		over_trends_days = trends_anlysis(rsi_si,oversold,1)["trends_days"]
		trend_dict['trends_days'] = f'[{over_trends_days}일]{trend_dict["trends_days"]}'
	elif rsi_current < standard :
		trend = '매도'
	
	# return : 추세전환(음→양)/[과]매수([Y일]X일)
	result = f'추세전환({trend_dict["transition"]})/{trend}({trend_dict["trends_days"]}일)'

	if dist_from_52w_high > 0 and rsi_current < trend_dict['high'] :
		result =  f'{result}/상승약화(하락가능)'
	if dist_from_52w_low < 0 and rsi_current > trend_dict['low'] :
		result =  f'{result}/하락약화(상승가능)'

	return result

# return = {'macd_trends' : macd_trends, 'osc_trends' : osc_trends, 'osc_trends_days' : osc_trends_days}
def get_macd_trends(price_si, short=12, long=26, signal=9, days=180, acceptable_transition_day = 3):
	standard = 0
	price_si = price_si[-days:]
	macd_si = price_si.ewm(span=short, min_periods=short-1, adjust=False).mean() - price_si.ewm(span=long, min_periods=long-1, adjust=False).mean()
	macd_signal_si = macd_si.ewm(span=signal, min_periods=signal-1, adjust=False).mean()
	macd =  macd_si.iloc[0]

	osc_si = macd_si - macd_signal_si
	osc_si = osc_si.dropna()
	osc_si = osc_si.iloc[::-1] # 행 순서 반대로
	osc_current = osc_si[0]
	trend_dict = trends_anlysis(osc_si, standard, acceptable_transition_day)

	if macd >= standard :
		macd_trends = '강세'
		if osc_current >= standard :
			osc_trends = '장기상승'
		else :
			osc_trends= '단기하락'

		if osc_current >= trend_dict['high'] :
			position = '최고점'
		else :
			position = '추락'
	else :
		macd_trends = '약세'
		if osc_current >= standard :
			osc_trends = '단기상승'
		else :
			osc_trends = '장기하락'

		if osc_current <= trend_dict['low'] :
			position = '최저점'
		else :
			position = '돌파'

	if osc_si[1] < 0 and osc_current > 0 :
		cross = '/GoldCross(buy)'
	elif osc_si[1] > 0 and osc_current < 0 :
		cross = '/DeadCross(sell)'
	else :
		cross = ''
	
	# return : 추세전환(음→양)/장기상승(X일)/최고점/
	result = f'추세전환({trend_dict["transition"]})/{osc_trends}({trend_dict["trends_days"]}일)/{position}{cross}'

	return {'macd_trends' : macd_trends, 'osc_trends' : result}

def get_stock_data(result):
	ticker = result['ticker']
	fd_y = finpie.Fundamentals(ticker,'yahoo') # yahoo finance 에서 들고온 데이터
	fd_f = finpie.Fundamentals(ticker)         # Finviz 에서 들어온 데이터
	stock_datas_f = fd_f.key_metrics().iloc[0]
	stock_datas_y = fd_y.key_metrics().iloc[0]

	revenue_estimate = fd_y.revenue_estimates()['avg_estimate']
	earnings_estimate = fd_y.earnings_estimate()['avg_estimate']
	shares_outstanding = stock_datas_y['shares_outstanding']
	rps_ttm = revenue_estimate.iloc[2] / shares_outstanding
	rps_after_2yr = revenue_estimate.iloc[3] / shares_outstanding
	eps_ttm = earnings_estimate.iloc[2]
	eps_after_2yr = earnings_estimate.iloc[3]

	ps_ratio_ttm = stock_datas_f['p_to_s']
	pe_ratio_ttm = stock_datas_f['p_to_e']
	if ps_ratio_ttm != '-':
		result['ps_ratio_ttm'] = ps_ratio_ttm
	if pe_ratio_ttm != '-':
		result['pe_ratio_ttm'] = pe_ratio_ttm

	ps_ratio_avg_5yr = 0
	pe_ratio_avg_5yr = 0
	com_name = tickers.company_name[ticker]
	df = pd.read_html(f'https://www.macrotrends.net/stocks/charts/{ticker}/{com_name}/price-sales', header=1)[0]
	if ~df.empty:
		df.fillna(0.0, inplace=True)
		df['Date']=df['Date'].str.extract('(\d{4})').astype(int)
		past_5_yr = df.iloc[0]["Date"]-5
		ps_ratio_avg_5yr= df.loc[df['Date'] > past_5_yr, "Price to Sales Ratio"].mean()
	
	df = pd.read_html(f'https://www.macrotrends.net/stocks/charts/{ticker}/{com_name}/pe-ratio', header=1)[0]
	if ~df.empty:
		df.fillna(0.0, inplace=True)
		df['Date']=df['Date'].str.extract('(\d{4})').astype(int)
		past_5_yr = df.iloc[0]["Date"]-5
		pe_ratio_avg_5yr = df.loc[df['Date'] > past_5_yr, "PE Ratio"].mean()

	result['dist_from_52w_high'] = round(stock_datas_f['52w_high']*100,2)
	result['dist_from_52w_low'] = round(stock_datas_f['52w_low']*100,2)
	
	# result['rsi_14']      = stock_datas_f['rsi_(14)']
	try:
		price_df = get_historical_price_data(ticker)
		#price_df = finpie.price_data.historical_prices(ticker)[-360:]
		price_si = price_df['Adj Close']
		result['adj_close']   = round(price_si.iloc[-1],2)  # stock_datas_f['prev_close']
		result['rsi_trends']  = get_rsi_trends(price_si, result['dist_from_52w_high'], result['dist_from_52w_low'], days=180, acceptable_transition_day=2)
		macd_trends = get_macd_trends(price_si, short=12, long=26, signal=9, days=180, acceptable_transition_day = 3)
		result['macd_trends'] = macd_trends['macd_trends']
		result['osc_trends']  = macd_trends['osc_trends']
	except:
		print('[Error] get_historical_price_data 가져오지 못함')
		pass
		
	if ps_ratio_avg_5yr != 0 and ps_ratio_avg_5yr != np.inf and ps_ratio_avg_5yr != -np.inf:
		result['stock_estimate_2yr_by_psr'] = round(rps_after_2yr*ps_ratio_avg_5yr,2)
	if pe_ratio_avg_5yr != 0 and pe_ratio_avg_5yr != np.inf and pe_ratio_avg_5yr != -np.inf:
		result['stock_estimate_2yr_by_per'] = round(eps_after_2yr*pe_ratio_avg_5yr,2)

	result.fillna(0.0, inplace=True)
	return result


# df 받아서 워크시트에 출력
def convert_df_to_ws(df, ws):
	for r in dataframe_to_rows(df, index=False, header=True):
		ws.append(r)
	return ws

# 워크시트 받아서 df로 반환
def convert_ws_to_df(ws, index_include, column_include):
	data = ws.values
	if index_include and column_include:
		cols = next(data)[0:]
		data = list(data)
		idx = [r[0] for r in data]
		data = (islice(r, 1, None) for r in data)
		df = pd.DataFrame(data, index=idx, columns=cols)
	elif index_include:
		data = list(data)
		idx = [r[0] for r in data]
		data = (islice(r, 1, None) for r in data)
		df = pd.DataFrame(data, index=idx, columns=None)
	elif column_include:
		cols = next(data)[0:]
		df = pd.DataFrame(data, index=None, columns=cols)
	else:
		df = pd.DataFrame(data)
	return df

def write_excel_report():
	# Portfolio Excel 파일 로드
	portfolio_wb = openpyxl.load_workbook(excel_path + portfolio_file_name)
	portfolio_ws = portfolio_wb[portfolio_file_sheet_title]
	portfolio_df = convert_ws_to_df(portfolio_ws, index_include=False, column_include=True)

	# 데이터 수집
	for index, row in portfolio_df.iterrows():
		print(row['ticker'])
		portfolio_df.loc[index] = get_stock_data(row)

	# 레포트 Excel 파일 생성
	report_wb = openpyxl.Workbook()
	report_ws = report_wb.active
	report_ws.title = report_file_sheet_title
	convert_df_to_ws(portfolio_df, report_ws)
	try :
		report_wb.save(excel_path + report_file_name)
	except FileNotFoundError:
		print('[Error] Excel 파일 찾을 수 없음. 경로 확인 바람')
	except PermissionError:
		print(f'[Error] Excel 파일 열려있거나 같은 파일명 존재. 새 파일 생성({new_report_file_name})')
		report_wb.save(excel_path + new_report_file_name)

	#original_df = pd.read_excel(excel_path, sheet_name=excel_sheet_name)
	#for index, row in original_df.iterrows():
	#	original_df.loc[index] = getStockData(row)

	#original_df.fillna(0.0, inplace=True)
	#original_df.to_excel(excel_path,sheet_name=excel_sheet_name, index=False)

# Main
write_excel_report()
