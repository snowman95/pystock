import finpie
import pandas as pd
import numpy as np
import openpyxl
import datetime
from openpyxl.chart import LineChart, BarChart, Reference
pd.set_option('display.max_row', 1000000)
pd.set_option('display.max_columns', 1000000)

billion = 1000000000.0 # 10억(B)
excel_path = 'C:\\Users\\hoon\\Desktop\\Stock\\Portfolio.xlsx'
excel_sheet_name = 'Portfolio'

from openpyxl.chart import (
    BarChart,
    StockChart,
    Reference,
    Series,
)
from openpyxl.chart.axis import DateAxis, ChartLines
from openpyxl.chart.updown_bars import UpDownBars

# 엑셀 작성 함수
def test(filename, sheet_title):
    wb = openpyxl.Workbook(filename)
    ws = wb.active
    ws.title = sheet_title

    # append 함수로 각 셀마다 데이터 입력
    ws.append(['이름', '국어', '영어'])
    ws.append(['김철수', 40, 60])
    ws.append(['김영희', 50, 50])
    ws.append(['임꺽정', 60, 80])

    ws2 = wb.create_sheet(title="Pi")
    # 행/열 하이라이트
    for cell in ws['A'] + ws[1]:
    	cell.style = 'Pandas'

    wb.save(filename)


def draw_chart(ws):
	# High-low-close
	c1 = StockChart()
	labels = Reference(ws, min_col=1, min_row=2, max_row=6)
	data = Reference(ws, min_col=4, max_col=6, min_row=1, max_row=6)
	c1.add_data(data, titles_from_data=True)
	c1.set_categories(labels)
	for s in c1.series:
	    s.graphicalProperties.line.noFill = True
	# marker for close
	s.marker.symbol = "dot"
	s.marker.size = 5
	c1.title = "High-low-close"
	c1.hiLowLines = ChartLines()
	

def draw_bar(ws):
	# 차트 초기화
	chart = BarChart()
	# chart = LineChart()
	
	# 데이터 영역 position 형태로 추가: (1,2) -> (4,3) 까지
	chartData = Reference(ws, min_col=2, max_col=3,
	                        min_row=1, max_row=4)
	
	# 카테고리(이름)에 해당하는 영역 position 형태로 추가
	category = Reference(ws, min_col=1, min_row=2, max_row=4)
	
	# 차트에 데이터 바인딩, title_from_data는 범례 값
	chart.add_data(chartData, titles_from_data=True)
	
	# 차트에 카테고리 바인딩
	chart.set_categories(category)
	
	# F1 셀에 차트 추가
	ws.add_chart(chart, 'F1')
